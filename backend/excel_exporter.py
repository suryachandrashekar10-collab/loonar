"""
Deviation register Excel export using openpyxl.
Generates a properly formatted, color-coded workbook ready for customer submission.
"""
import io
from datetime import datetime


STATUS_COLORS = {
    "pending":       ("FFF9C4", "F9A825"),   # amber bg, amber border
    "accepted":      ("C8E6C9", "388E3C"),   # green
    "rejected":      ("FFCDD2", "C62828"),   # red
    "clarification": ("BBDEFB", "1565C0"),   # blue
}

RISK_COLORS = {
    "high":   "FFCDD2",
    "medium": "FFF9C4",
    "low":    "F1F8E9",
}


def generate_deviation_register(
    deviations: list[dict],
    project: dict,
) -> bytes:
    """
    Returns Excel file bytes for the deviation register.
    Call with deviations from API + project metadata.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Deviation Register"

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"DEVIATION REGISTER — {project.get('name', 'Project').upper()}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A237E")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    meta = [
        ("Customer:", project.get("customer", "")),
        ("RFQ Ref:", project.get("rfq_ref", "")),
        ("Value:", f"€{project.get('value_eur', 0):,.0f}" if project.get("value_eur") else ""),
        ("Exported:", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row_i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=row_i, column=1, value=label).font = Font(bold=True, color="555555")
        ws.cell(row=row_i, column=2, value=value)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    # ── Header row ─────────────────────────────────────────────────────────────
    headers = [
        "Dev ID", "Clause / Ref", "Doc Reference",
        "Customer Specification (Verbatim)",
        "Proposed Deviation / Exception",
        "Technical Justification",
        "Status",
    ]
    header_row = 7
    header_fill = PatternFill("solid", fgColor="37474F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 28
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_i, dev in enumerate(deviations, start=header_row + 1):
        status = dev.get("status", "pending")
        bg_color, _ = STATUS_COLORS.get(status, ("FFFFFF", "AAAAAA"))

        row_data = [
            dev.get("dev_id", ""),
            dev.get("clause", ""),
            dev.get("doc_ref", ""),
            dev.get("customer_spec", ""),
            dev.get("proposed_deviation", ""),
            dev.get("justification", ""),
            status.upper(),
        ]

        row_fill = PatternFill("solid", fgColor=bg_color)
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col, value=value)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 7:  # Status
                status_fill_color, status_font_color = STATUS_COLORS.get(status, ("FFFFFF", "000000"))
                cell.fill = PatternFill("solid", fgColor=status_fill_color)
                cell.font = Font(bold=True, color=status_font_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row height based on longest cell
        max_len = max(len(str(v or "")) for v in row_data)
        ws.row_dimensions[row_i].height = max(24, min(80, max_len // 3))

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [10, 14, 20, 50, 50, 50, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Summary tab ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Status"
    ws2["B1"] = "Count"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)

    status_counts: dict[str, int] = {}
    for dev in deviations:
        s = dev.get("status", "pending")
        status_counts[s] = status_counts.get(s, 0) + 1

    for i, (s, count) in enumerate(status_counts.items(), start=2):
        ws2.cell(row=i, column=1, value=s.title())
        ws2.cell(row=i, column=2, value=count)

    ws2.cell(row=len(status_counts) + 3, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=len(status_counts) + 3, column=2, value=len(deviations)).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
